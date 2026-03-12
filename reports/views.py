import csv
import io
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.db.models import Count, Avg, Q
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from accounts.models import User, Department
from courses.models import Course, Enrollment
from grades.models import SemesterResult

@login_required
def reports_dashboard(request):
    if not request.user.is_super_admin:
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    
    context = {
        'total_students': User.objects.filter(role='student').count(),
        'total_courses': Course.objects.count(),
        'total_departments': Department.objects.count(),
        'total_admins': User.objects.filter(role__in=['sub_admin','super_admin']).count(),
        'departments': Department.objects.annotate(
            student_count=Count('user', filter=Q(user__role='student')),
            course_count=Count('course')
        ),
    }
    return render(request, 'reports/dashboard.html', context)

@login_required
def export_students_excel(request):
    if not request.user.is_super_admin:
        return HttpResponse('Access denied', status=403)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"
    
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    headers = ['Student ID','First Name','Last Name','Username','Email','Department','Gender','Date of Birth','Phone','Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    students = User.objects.filter(role='student').select_related('department')
    for row, student in enumerate(students, 2):
        data = [
            student.student_id or '', student.first_name, student.last_name,
            student.username, student.email,
            student.department.name if student.department else '',
            student.get_gender_display() if student.gender else '',
            str(student.date_of_birth) if student.date_of_birth else '',
            student.phone, 'Active' if student.is_active else 'Inactive'
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            if row % 2 == 0:
                cell.fill = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")
    
    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 30)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="students_report.xlsx"'
    wb.save(response)
    return response

@login_required
def export_grades_excel(request):
    if not request.user.is_super_admin:
        return HttpResponse('Access denied', status=403)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Semester Results"
    
    header_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    headers = ['Student ID','Student Name','Course Code','Course Name','Department','Semester','Score','Grade','GPA Points','Published']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    results = SemesterResult.objects.all().select_related('student','course','semester','course__department')
    for row, result in enumerate(results, 2):
        data = [
            result.student.student_id or '',
            result.student.get_full_name(),
            result.course.code,
            result.course.name,
            result.course.department.name if result.course.department else '',
            str(result.semester),
            result.total_score,
            result.letter_grade,
            result.gpa_points,
            'Yes' if result.published else 'No'
        ]
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)
    
    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 30)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="grades_report.xlsx"'
    wb.save(response)
    return response

@login_required
def export_students_csv(request):
    if not request.user.is_super_admin:
        return HttpResponse('Access denied', status=403)
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="students.csv"'
    writer = csv.writer(response)
    writer.writerow(['Student ID','First Name','Last Name','Email','Department','Status'])
    for s in User.objects.filter(role='student').select_related('department'):
        writer.writerow([s.student_id, s.first_name, s.last_name, s.email, s.department.name if s.department else '', 'Active' if s.is_active else 'Inactive'])
    return response

@login_required
def import_students(request):
    if not request.user.is_super_admin:
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    if request.method == 'POST' and request.FILES.get('file'):
        f = request.FILES['file']
        decoded = f.read().decode('utf-8')
        reader = csv.DictReader(io.StringIO(decoded))
        created = 0
        for row in reader:
            if not User.objects.filter(username=row.get('username','')).exists():
                dept = None
                if row.get('department'):
                    dept, _ = Department.objects.get_or_create(name=row['department'], defaults={'code': row['department'][:10]})
                user = User(
                    username=row.get('username',''),
                    first_name=row.get('first_name',''),
                    last_name=row.get('last_name',''),
                    email=row.get('email',''),
                    role='student',
                    department=dept,
                    student_id=row.get('student_id','') or None,
                )
                user.set_password('School@123')
                user.save()
                created += 1
        messages.success(request, f'Imported {created} students. Default password: School@123')
        return redirect('user_list')
    return render(request, 'reports/import_students.html')

@login_required
def enrollment_report(request):
    if not request.user.is_super_admin:
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    courses = Course.objects.annotate(enrollment_count=Count('enrollments',filter=Q(enrollments__status='enrolled'))).select_related('department','semester')
    return render(request, 'reports/enrollment_report.html', {'courses': courses})

@login_required
def gpa_report(request):
    if not request.user.is_super_admin:
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    students = User.objects.filter(role='student').prefetch_related('semester_results')
    student_gpas = []
    for s in students:
        results = s.semester_results.filter(published=True)
        if results.exists():
            avg_gpa = sum(r.gpa_points for r in results) / results.count()
            student_gpas.append({'student': s, 'gpa': round(avg_gpa, 2), 'courses': results.count()})
    student_gpas.sort(key=lambda x: x['gpa'], reverse=True)
    return render(request, 'reports/gpa_report.html', {'student_gpas': student_gpas})

from accounts.models import Department as Dept
